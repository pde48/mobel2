# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

from odoo import api, fields, models, _, tools
from odoo.osv import expression
from datetime import datetime, time

from odoo.exceptions import UserError, ValidationError
import xlwt
import io
import os.path
import base64
import openpyxl
from io import BytesIO

#EXPORT
class ReportExportManagment(models.TransientModel):
    _name = 'report.export.managment'
    _description = 'ReportExportManagment'

    name = fields.Char(string='Name',size=64,required=True,)
    date_order_line = fields.Datetime(string='Order Date Line', required=True, readonly=True, index=True, copy=False, default=fields.Datetime.now)

    report_export_managment_ids = fields.One2many(
        'report.export.list.managment',
        'report_export_id',
        string='ReportExportListManagment',
    )

#EXPORT
class ReportExportListManagment(models.TransientModel):
    _name = 'report.export.list.managment'
    _description = 'ReportExportListManagment'


    #product_id = fields.Char(string='# Product ID',required=True,)
    product_id = fields.Many2one('product.product', string='Product')
    import_id = fields.Many2one('purchase.import',string='Purchase Import')
    product_qty = fields.Float(string='Product Qty',required=True,)    

    product_default_code = fields.Char(string='Product Default Code',required=False,)
    url_product = fields.Char(string="URL",store=True)

    

    report_export_id = fields.Many2one(
        'report.export.managment',
        string='Report Managament',
    )



#IMPORT
class ReportImportManagment(models.TransientModel):
    _name = 'report.import.managment'
    _description = 'ReportImportManagment'

    name = fields.Char(string='Name',size=64,required=False,)
    code_supplier = fields.Char(string='code_supplier',size=64,required=False,)
    supplier = fields.Char(string='supplier',size=64,required=False,)
    date_order_line = fields.Datetime(string='Order Date Line', required=True, readonly=True, index=True, copy=False, default=
        fields.Datetime.now)

    report_import_managment_ids = fields.One2many(
        'report.import.list.managment',
        'report_import_id',
        string='ReportImportListManagment',
    )


#IMPORT
class ReportImportListManagment(models.TransientModel):
    _name = 'report.import.list.managment'
    _description = 'ReportImportListManagment'


    product = fields.Char(string='# Product ID',required=True,)
    product_id = fields.Many2one('product.product', string='Product')
    import_id = fields.Many2one('purchase.import',string='Purchase Import')
    import_name = fields.Char(string='Purchase Import')
    product_qty = fields.Float(string='Product Qty',required=True,)    

    product_default_code = fields.Char(string='Product Default Code',required=False,)
    url_product = fields.Char(string="URL",store=True)
    

    report_import_id = fields.Many2one(
        'report.import.managment',
        string='ReportImportListManagment',
    )




class OperationsPurchasesFiles(models.TransientModel):
    _name = "operations.purchases.files"
    _description = "Operations Purchases Files"


    @api.model
    def _count(self):
        return len(self._context.get('active_ids', []))

    count = fields.Integer(default=_count, string='Cantidad de Ordenes Seleccionadas')
    decision_operations = fields.Selection(selection=[
        ('Export', 'Export'),
        ('Import', 'Import'),

    ], string='Operation', copy=False,required=True )

    company_id = fields.Many2one('res.company', 'Company', required=True, index=True, default=lambda self: self.env.company.id)

    file = fields.Binary(string="File", required=True)

    def add_set_line_report(self,report_id,import_id,line_purchase_id,product_id,url_product,default_code):
        print("import")
        print(import_id)
        res_search_1 = self.env['purchase.order.line'].search_read([('id','=',line_purchase_id)],['product_id','product_qty',])
        print(res_search_1)
        for rec_po_line in res_search_1:

            product_qty_ava = rec_po_line['product_qty']
            #verificar que si existe en modelo de lineas de report
            res_validate_line = self.env['report.export.list.managment'].search_read(['&','&',('product_id','=',product_id),('import_id','=',import_id),('report_export_id','=',report_id)],['id','product_qty'])

            if(res_validate_line):
                print("EXISTE")
                #obtener la cantidad y sumar la cantidad que viene
                for line_data in res_validate_line:
                    id_line_report = line_data['id']
                    print('product_qty')
                    print(line_data['product_qty'])
                    print(product_qty_ava)
                    qty_update = float(line_data['product_qty']) + float(product_qty_ava)

                res_up_line = self.env['report.export.list.managment'].search([('id','=',id_line_report)])
                res_up_line.write({'product_qty': float(qty_update)})

            else:
                print("NO EXISTE")
                #sino existe crear linea de report
                search_df_code = self.env['product.product'].search_read([('id','=',product_id)],['default_code','url_product'],limit=1)
                for data_res3 in search_df_code:
                    default_code2 = data_res3['default_code']
                    url_product2 = data_res3['url_product']

                rec_create_line = self.env['report.export.list.managment']
                vals_create_line = {
                                   
                                     'product_id': product_id,
                                     'import_id': import_id,
                                     'url_product': url_product2,
                                     'product_qty': product_qty_ava,
                                     'product_default_code':default_code2,
                                     'report_export_id':report_id,
                }
                rec_create_line.create(vals_create_line)





    def download_file_operation(self):
            if(self.decision_operations == 'Export'):
                #Create Report 
                vals_create = {'name': 'Report Export '+str(datetime.now())}
                res_create = self.env['report.export.managment'].create(vals_create)
                report_id = res_create.id

                search_product1 = self.env['purchase.order.line'].browse(self._context.get('active_ids', []))
                for rec_2 in search_product1:
                    line_purchase_id = rec_2.id
                    import_id = rec_2.mapped('import_id').id
                    product_id = rec_2.mapped('product_id').id
                    url_product = rec_2.mapped('url_product')
                    default_code = rec_2.mapped('default_code')                    

                    set_line_report = self.add_set_line_report(report_id,import_id,line_purchase_id,product_id,url_product,default_code)

                """Download"""
                row_index = 4
                workbook = xlwt.Workbook(encoding="utf-8")
                style_title = xlwt.easyxf(
                   "font:height 200; font: name Liberation Sans, bold on,color black; align: horiz left")
                currency = xlwt.easyxf('font: height 180; align: wrap yes, horiz right', num_format_str='$#0')
                percent = xlwt.easyxf('font: height 180; align: wrap yes, horiz right', num_format_str='#0%')
                budget_name = "Reporte_11"
                budget_name2 = "Reporte_1"
                today = datetime.today().strftime("%d-%m-%Y")
                worksheet = workbook.add_sheet(budget_name)
                j = 0
                worksheet.write_merge(0, 0, 0, 0, 'Proveedor', style_title);j += 1
                worksheet.write_merge(1, 1, 0, 0, 'Numero de Orden', style_title);j += 1

                k = 4;
                j = 0

                """cabecera"""

                #Nombre del producto Código  Precio unidad   Precio total
                #worksheet.write_merge(k, k, j, j, 'Id', style_title);j += 1
                worksheet.write_merge(k, k, j, j, '#Compra', style_title);j += 1
                worksheet.write_merge(k, k, j, j, '#Store', style_title);j += 1
                worksheet.write_merge(k, k, j, j, '#Codigo', style_title);j += 1
                worksheet.write_merge(k, k, j, j, '#Link URL', style_title);j += 1                
                worksheet.write_merge(k, k, j, j, '#Cantidad', style_title);j += 1
                """línea"""

                #purchases_orders_lines = self.env['purchase.order.line'].browse(self._context.get('active_ids', []))
                report_managment_line = self.env['report.export.list.managment'].search_read([('report_export_id','=',report_id)])                
                for data_exp in report_managment_line:
                    print(data_exp)
                    #print(rec_.mapped('import_id'))
                    res_search_import_00 = self.env['purchase.import'].search_read([('id','=',data_exp['import_id'][0])],['name'] ,limit=1)
                    for data_rec2 in res_search_import_00:
                        import_name_res = data_rec2['name']

                    j = 0
                    row_index += 1
                    #worksheet.write_merge(row_index,row_index,j, j, str(data_exp['id']), );j += 1
                    worksheet.write_merge(row_index,row_index,j, j, str(import_name_res), );j += 1
                    worksheet.write_merge(row_index,row_index,j, j, str('STORE'), );j += 1
                    worksheet.write_merge(row_index,row_index,j, j, str(data_exp['product_default_code']), );j += 1
                    worksheet.write_merge(row_index,row_index,j, j, str(data_exp['url_product']), );j += 1
                    worksheet.write_merge(row_index,row_index,j, j, str(data_exp['product_qty']), );j += 1
                    #worksheet.write_merge(row_index,row_index,j, j, str(data_exp['product_id']), );j += 1



                worksheet.col(0).width = 6000
                worksheet.col(1).width = 8000
                worksheet.col(2).width = 10000
                worksheet.col(3).width = 30000
                fp = io.BytesIO()
                workbook.save(fp)
                fp.seek(0)
                data = fp.read()
                fp.close()
                data_b64 = base64.encodestring(data)

                doc = self.env['ir.attachment'].create({
                   'name': '%s.xls' % (budget_name2),
                   'datas': data_b64,
                   'name': '%s.xls' % (budget_name2),
                })

                return {
                   'type': "ir.actions.act_url",
                   'url': "web/content/?model=ir.attachment&id=" + str(
                       doc.id) + "&filename_field=datas_fname&field=datas&download=true&filename=" + str('lista_lineas_compras-'),
                   'target': "self",
                   'no_destroy': False,
               }
            else:
                if(self.decision_operations == 'Import'):
                    #import
                    try:
                        wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.file)), read_only=False  )
                        ws = wb.active
                        #data proveedor y numero               
                        for record in ws.iter_rows(min_row=0, max_row=1, min_col=0,max_col=2, values_only=True):
                            supplier = record[1]

                        for record_0 in ws.iter_rows(min_row=2, max_row=2, min_col=0,max_col=2, values_only=True):
                            code_supplier = record_0[1]

                        vals_create_rc_2 = {'name': 'IMPORT - '+str(datetime.now()),'supplier':supplier,'code_supplier':code_supplier}
                        res_created_rec_2 = self.env['report.import.managment'].create(vals_create_rc_2)

                        #codigo
                        for record3 in ws.iter_rows(min_row=6, max_row=None, min_col=1,max_col=5, values_only=True):
                            rec_create_line2 = self.env['report.import.list.managment']
                            vals_create_line2 = {
                                                 'product': record3[2],
                                                 'import_name': record3[0],
                                                 'product_qty': record3[4],
                                                 'url_product': record3[3],
                                                 'report_import_id':res_created_rec_2.id,
                                                }
                            rec_create_line2.create(vals_create_line2)



                        #search id 
                        res_search_00 = self.env['report.import.list.managment'].search_read([('report_import_id','=',res_created_rec_2.id)],['import_name','product','url_product','product_qty'])
                        for data_res in res_search_00:
                            #print(data_res)
                            import_name = data_res['import_name']
                            product = data_res['product']
                            url_product = data_res['url_product']
                            product_qty = data_res['product_qty']

                            res_search_import_00 = self.env['purchase.import'].search_read([('name','ilike',import_name)],limit=1)
                            for data_rec2 in res_search_import_00:
                                import_id_res = data_rec2['id']

                            search_df_code = self.env['product.product'].search_read([('default_code','=',product)],['default_code','url_product'],limit=1)
                            for data_res3 in search_df_code:
                                id_product = data_res3['id']

                            res_search_po_line_00 = self.env['purchase.order.line'].search_read(['&',('import_id','=',import_id_res),('product_id','=',id_product)],['product_id','number_puchase_supplier','product_qty','date_order_line'],order="date_order_line desc")
                            for data_res2 in res_search_po_line_00:
                                for upt in range(int(product_qty)):
                                    res_update= self.env['purchase.order.line'].search([('id','=',data_res2['id'])])
                                    res_update.write({'purchase_ok':True})

                            
                        #search if the customer exist else create
                    except:
                        raise UserError(_('Please insert a valid file'))
            
            return {'type': 'ir.actions.act_window_close'}
